#!/usr/bin/env python
"""This program inputs a MemberHub directory dump, and analyzes it.
"""
import family
import roster
import os
from openpyxl import load_workbook

NUM_ROSTER_FIELDS = 5

def ReadRosterAdultsFromMostRecent():
    """ roster_tools.ReadRosterAdultsFromMostRecent
    PURPOSE:
    Generates a list of adult names in the newest roster file.
    INPUT:
    - none
    OUTPUTS:
    - adults_list -- list of adult name fields in the newest roster file.
    ASSUMPTIONS:
    - none
    """
    ##
    ## Find the files in the "Roster" folder with ".xlsx" extension, sort them by
    ## date, and pick the most recently added
    file_path = os.path.abspath("./Roster/")
    with os.scandir(file_path) as raw_files:
        files = [file for file in raw_files \
                    if not(file.name.startswith('~')) and (file.name.endswith('.xlsx'))]
        files.sort(key=lambda x: os.stat(x).st_mtime, reverse=True)
        file_name = file_path + "/" +files[0].name
    
    ##
    ## Load the workbook, and select the active/only worksheet
    wb = load_workbook(file_name)
    ws = wb.active
    ##
    ## Copy all the values in column 'D' for all rows beyond the title row into
    ## the output list
    adults_list = []
    for fields in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        adults_list.append(fields[0].value)
        
    return adults_list


def ReadRosterFromFile(file_name, hub_map, rosterC):
    """ roster_tools.ReadRosterFromFile
    PURPOSE:
    Reads a roster file with the following fields:
    <**Last Name>,<**First Name>,<**Grade>,<**Parent/Guardian Name(s)>,<***Teacher Name>
    **  - indicates always required field
    *** - indicates field that is required when Grade field is < 6
    INPUT:
    - file_name -- name of the roster file
    - hub_map   -- dictionary that maps hub names to hub IDs
    - rosterC   -- the Roster object containing the errata
    OUTPUTS:
    - roster    -- list of families extracted from the roster
    ASSUMPTIONS:
    1. First row of the file is the column headers...not a member of the roster.
    """
    wb            = load_workbook(file_name)
    ws            = wb.active
    student_count = -1

    for fields in ws.values:

        ## Skip the first row
        if student_count < 0:
            student_count = 0
            continue

        ## Skip any row for which all fields are not populated
        empty_field_found = False
        for field in fields:
            if field == None or field == "":
                empty_field_found = True
                print("Found row with missing required fields:", fields)
                continue
        if empty_field_found:
            continue

        ## each row represents one student
        student_count += 1

        ## treat the student as a member of a new family...for now
        new_family = family.RosterFamily(adults_raw_name=fields[3])
        new_family.AddToFamily(child_first  = fields[1],
                               child_last   = fields[0],
                               grade        = fields[2],
                               adult_names  = fields[3],
                               teacher_name = fields[4],
                               hub_map      = hub_map,
                               rosterC      = rosterC)

        # if new_family is the same as a family already in the roster, then combine
        # families.  Otherwise, append new_family at the end of the roster.
        for roster_entry in rosterC.GetRoster():
            if roster_entry.IsSameFamily(new_family):
                roster_entry.CombineWith(new_family)
                break
        else:
            rosterC.append(new_family)

    print("%d students processed %d families." % (student_count, len(rosterC)))

    return rosterC.GetRoster()


def GetRosterFileName():
    """ roster_tools.GetRosterFileName
    PURPOSE:
    Gives the user a list of possible roster files, and processes their selection.
    INPUTS:
    None
    OUTPUTS:
    - file_name - the selected roster file name
    ASSUMPTIONS:
    - Assumes the candidate roster files are stored in a subfolder called 'Roster'
    """
    print ("These are the potential roster files:")
    file_path = os.path.abspath("./Roster/")
    with os.scandir(file_path) as raw_files:
        files = [file for file in raw_files \
                    if not(file.name.startswith('~')) and (file.name.endswith('.xlsx'))]
        files.sort(key=lambda x: os.stat(x).st_mtime, reverse=True)

        max_index = 0
        file_number = 1
        while int(file_number) >= max_index:
            for file in files:
                max_index += 1
                print("%d) %s" % (max_index, file.name))

            file_number = input("Enter list number of file or press <enter> to use '" + files[0].name + "':")
            if not file_number:
                return file_path + "/" +files[0].name
            elif 0 < int(file_number) and int(file_number) <= max_index:
                return file_path + "/" + files[int(file_number)-1].name
            else:
                max_index = 0
                print("The selection made is out of range.  Please try again.")

def ReadRoster(hub_map):
    """ roster_tools.ReadRoster
    PURPOSE:
    Prompts the user for roster file name and proceeds to read the file.
    INPUT:
    - hub_map   -- mapping of teacher names to hub numbers
    OUTPUTS:
    - roster    -- list of families extracted from the roster
    ASSUMPTIONS:
    - All the candidate rosters reside in a folder called "Roster" under the
      run directory.
    - All candidate rosters are Microsoft Excel files.
    """
    return ReadRosterFromFile(GetRosterFileName(), hub_map, roster.Roster())

